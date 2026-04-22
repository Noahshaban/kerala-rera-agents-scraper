"""
Kerala RERA Agents — Full Dataset Scraper (Pages 1 to 45)
==========================================================
Scrapes all agents from:
    https://rera.kerala.gov.in/agents          (page 1)
    https://rera.kerala.gov.in/agents?page=2   (page 2)
    ...
    https://rera.kerala.gov.in/agents?page=45  (page 45)

This approach is more reliable than clicking the Next button because
it navigates to each page directly by URL — no button-detection needed.

Install dependencies (run once):
    pip install playwright pandas openpyxl
    playwright install chromium

Run:
    python kerala_rera_full_scraper.py
"""

import re
import sys
import time
import logging
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PWTimeoutError
from playwright.sync_api import sync_playwright


# ──────────────────────────────────────────────────────────────
# LOGGING  (timestamped console output)
# ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────
BASE_URL        = "https://rera.kerala.gov.in/agents"
FIRST_PAGE      = 1       # start page (page 1 has no ?page= parameter)
LAST_PAGE       = 45      # last page: https://rera.kerala.gov.in/agents?page=45
OUTPUT_FILE     = r"/home/miracle/Downloads/Kerala_RERA_Full_Dataset.xlsx"
SHEET_NAME      = "All Agents"

PAGE_DELAY_SEC  = 3.0     # polite pause between page loads (seconds)
NAV_TIMEOUT_MS  = 90_000  # 90 s — maximum time to wait for a page to load
TABLE_WAIT_MS   = 25_000  # 25 s — maximum time to wait for table rows to appear

# CSS selector that confirms the data table has rendered on the page
ROW_SELECTOR    = "table tbody tr td"

# Raw columns in the order they appear left-to-right in the HTML table
RAW_COLUMNS = [
    "Agent Name",
    "Agent Type",
    "Address",
    "Land Mark",
    "Email ID",
    "Mobile No",
    "Certificate No",
]

# The four columns the client wants in the final Excel output
FINAL_COLUMNS = ["Name", "Mobile No", "Email ID", "Address"]


# ──────────────────────────────────────────────────────────────
# SECTION 1 — BUILD PAGE URLS
# ──────────────────────────────────────────────────────────────

def build_page_url(page_number: int) -> str:
    """
    Return the correct URL for a given page number.

    Page 1  ->  https://rera.kerala.gov.in/agents          (no query string)
    Page 2  ->  https://rera.kerala.gov.in/agents?page=2
    Page 45 ->  https://rera.kerala.gov.in/agents?page=45
    """
    if page_number == 1:
        return BASE_URL
    return f"{BASE_URL}?page={page_number}"


# ──────────────────────────────────────────────────────────────
# SECTION 2 — TABLE PARSING
# ──────────────────────────────────────────────────────────────

def parse_current_page(page) -> list[dict]:
    """
    Extract every agent row visible in the table on the currently loaded page.

    Steps:
      1. wait_for_selector blocks execution until at least one <td> is in the DOM,
         guaranteeing the table has fully rendered before we read it.
      2. All <tbody tr> rows are queried and their <td> cells read left-to-right.
      3. Rows shorter than RAW_COLUMNS are padded with empty strings so
         dictionary indexing never raises an IndexError.

    Returns a list of dicts with RAW_COLUMNS as keys.
    Raises PlaywrightTimeoutError if the table never appears within TABLE_WAIT_MS.
    """
    # Block until table data cells are present in the DOM
    page.wait_for_selector(ROW_SELECTOR, timeout=TABLE_WAIT_MS)

    rows_data: list[dict] = []
    table_rows = page.query_selector_all("table tbody tr")

    for row in table_rows:
        cells = row.query_selector_all("td")
        if not cells:
            continue   # skip rows that contain no <td> elements

        # Read each cell text and strip surrounding whitespace
        texts = [cell.inner_text().strip() for cell in cells]

        # Pad to at least len(RAW_COLUMNS) to avoid index-out-of-range errors
        if len(texts) < len(RAW_COLUMNS):
            texts += [""] * (len(RAW_COLUMNS) - len(texts))

        rows_data.append({
            "Agent Name":     texts[0],
            "Agent Type":     texts[1],
            "Address":        texts[2],
            "Land Mark":      texts[3],
            "Email ID":       texts[4],
            "Mobile No":      texts[5],
            "Certificate No": texts[6],
        })

    return rows_data


# ──────────────────────────────────────────────────────────────
# SECTION 3 — MAIN SCRAPING LOOP  (URL-based, page 1 to 45)
# ──────────────────────────────────────────────────────────────

def scrape_pages_1_to_45() -> pd.DataFrame:
    """
    Open a headless Chromium browser and iterate from page 1 to page 45
    by constructing the URL for each page directly — no button clicking needed.

    Error handling:
      - Per-page failures are caught and logged; the loop continues.
      - If 3 consecutive pages fail, the loop stops early and saves
        whatever has been collected so far.
      - Ctrl+C (KeyboardInterrupt) triggers a graceful early save.

    Returns a raw DataFrame with RAW_COLUMNS columns.
    """
    all_rows: list[dict]    = []
    consecutive_errors: int = 0
    MAX_CONSECUTIVE_ERRORS  = 3
    total_pages             = LAST_PAGE - FIRST_PAGE + 1   # 45

    # Track current page number outside the loop so KeyboardInterrupt can log it
    current_page_num = FIRST_PAGE

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
        )
        browser_page = context.new_page()

        log.info("Starting scrape: pages %d to %d  (%d pages total)",
                 FIRST_PAGE, LAST_PAGE, total_pages)
        log.info("Press Ctrl+C at any time to stop early and save collected data.")
        log.info("-" * 60)

        try:
            for page_num in range(FIRST_PAGE, LAST_PAGE + 1):
                current_page_num = page_num
                url      = build_page_url(page_num)
                progress = f"[{page_num:>2}/{LAST_PAGE}]"

                # ── Navigate directly to this page's URL ─────────────
                log.info("%s  Loading %s", progress, url)
                try:
                    browser_page.goto(
                        url,
                        wait_until="networkidle",
                        timeout=NAV_TIMEOUT_MS,
                    )
                except PWTimeoutError:
                    # networkidle timed out — page may still be usable; try to parse
                    log.warning(
                        "%s  networkidle timeout — will still attempt to parse the table.",
                        progress,
                    )
                except Exception as nav_err:
                    consecutive_errors += 1
                    log.warning(
                        "%s  Navigation error: %s  (streak: %d/%d)",
                        progress, nav_err,
                        consecutive_errors, MAX_CONSECUTIVE_ERRORS,
                    )
                    if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                        log.error("Too many consecutive navigation errors — stopping early.")
                        break
                    time.sleep(PAGE_DELAY_SEC)
                    continue

                # ── Extract table rows from the loaded page ───────────
                try:
                    rows = parse_current_page(browser_page)

                    if rows:
                        all_rows.extend(rows)
                        log.info(
                            "%s  Extracted %d rows  |  running total: %d",
                            progress, len(rows), len(all_rows),
                        )
                        consecutive_errors = 0   # success — reset the error streak
                    else:
                        log.warning("%s  Table loaded but contained 0 data rows.", progress)
                        consecutive_errors += 1

                except PWTimeoutError:
                    consecutive_errors += 1
                    log.warning(
                        "%s  Table did not appear within %d ms  (streak: %d/%d)",
                        progress, TABLE_WAIT_MS,
                        consecutive_errors, MAX_CONSECUTIVE_ERRORS,
                    )
                except Exception as parse_err:
                    consecutive_errors += 1
                    log.warning(
                        "%s  Parse error: %s  (streak: %d/%d)",
                        progress, parse_err,
                        consecutive_errors, MAX_CONSECUTIVE_ERRORS,
                    )

                # Stop if too many consecutive pages have failed
                if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                    log.error(
                        "%d consecutive page failures — stopping early to preserve collected data.",
                        MAX_CONSECUTIVE_ERRORS,
                    )
                    break

                # Polite delay before loading the next page
                if page_num < LAST_PAGE:
                    time.sleep(PAGE_DELAY_SEC)

        except KeyboardInterrupt:
            log.warning(
                "Interrupted by user at page %d — saving %d rows collected so far.",
                current_page_num, len(all_rows),
            )

        finally:
            browser.close()
            log.info("-" * 60)

    log.info("Scraping finished.  Total raw rows collected: %d", len(all_rows))

    if not all_rows:
        return pd.DataFrame(columns=RAW_COLUMNS)

    return pd.DataFrame(all_rows, columns=RAW_COLUMNS)


# ──────────────────────────────────────────────────────────────
# SECTION 4 — DATA CLEANING
# ──────────────────────────────────────────────────────────────

def clean_email(value: str) -> str:
    """
    Fix obfuscated email addresses commonly found on government sites:
      user[at]domain.com  ->  user@domain.com
      user at domain.com  ->  user@domain.com
    Also collapses multiple spaces and strips surrounding whitespace.
    """
    if not isinstance(value, str):
        return value
    value = re.sub(r"\[at\]",   "@", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+at\s+", "@", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+",      " ", value).strip()
    return value


def clean_address(value: str) -> str:
    """
    Normalize address text scraped from the HTML table:
      - Convert newlines and tab characters to a single space.
      - Collapse runs of multiple spaces into one space.
      - Strip surrounding whitespace.
    """
    if not isinstance(value, str):
        return value
    value = re.sub(r"[\t\r\n]+", " ", value)
    value = re.sub(r" {2,}",     " ", value)
    return value.strip()


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Full cleaning pipeline applied to the raw scraped DataFrame.

    Steps:
      1.  Rename 'Agent Name' -> 'Name'
      2.  Strip whitespace from all text columns
      3.  Fix obfuscated email addresses  ([at] -> @)
      4.  Normalize address whitespace / newlines
      5.  Replace empty strings with NaN (needed for correct dedup/dropna)
      6.  Drop fully duplicate rows
      7.  Drop rows where ALL four output columns are empty / NaN
      8.  Keep only the four client-requested columns
      9.  Fill remaining NaN with empty string (clean Excel appearance)
      10. Reset the integer index

    Returns a clean DataFrame with FINAL_COLUMNS only.
    """
    df = df.copy()

    # 1
    df.rename(columns={"Agent Name": "Name"}, inplace=True)

    # 2
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip()

    # 3
    df["Email ID"] = df["Email ID"].apply(clean_email)

    # 4
    df["Address"] = df["Address"].apply(clean_address)

    # 5
    df.replace("", pd.NA, inplace=True)

    # 6
    before = len(df)
    df.drop_duplicates(inplace=True)
    log.info("Deduplication removed %d duplicate row(s).", before - len(df))

    # 7
    df.dropna(subset=FINAL_COLUMNS, how="all", inplace=True)

    # 8
    df = df[FINAL_COLUMNS]

    # 9
    df.fillna("", inplace=True)

    # 10
    df.reset_index(drop=True, inplace=True)

    log.info("Clean dataset ready: %d rows x %d columns.", len(df), len(df.columns))
    return df


# ──────────────────────────────────────────────────────────────
# SECTION 5 — EXCEL EXPORT
# ──────────────────────────────────────────────────────────────

def export_to_excel(df: pd.DataFrame, filepath: str) -> None:
    """
    Save the clean DataFrame to a professionally formatted Excel workbook.

    Formatting applied:
      - Dark-blue bold header row with white Arial text
      - Alternating light-blue / white row fill for easy reading
      - Thin grey borders on every cell
      - Auto-fitted column widths (capped at 60 characters)
      - Frozen header row (stays visible while scrolling down)
      - Second sheet 'Scrape Info' with run metadata
    """
    # Write raw data via pandas (openpyxl engine = UTF-8 safe)
    df.to_excel(filepath, index=False, sheet_name=SHEET_NAME, engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb[SHEET_NAME]

    # Style definitions
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # deep navy blue
    ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")   # very light blue
    NO_FILL     = PatternFill()                            # transparent / no fill
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    DATA_FONT   = Font(name="Arial", size=10)
    H_ALIGN     = Alignment(horizontal="center", vertical="center")
    D_ALIGN     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Style the header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = H_ALIGN
        cell.border    = BORDER

    # Style data rows with alternating fill
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        row_fill = ALT_FILL if row_idx % 2 == 0 else NO_FILL
        for cell in row:
            cell.font      = DATA_FONT
            cell.alignment = D_ALIGN
            cell.border    = BORDER
            cell.fill      = row_fill

    # Auto-fit column widths (max 60 chars wide)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value),
            default=12,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Metadata sheet ────────────────────────────────────────
    summary = wb.create_sheet("Scrape Info")
    summary.sheet_properties.tabColor = "4472C4"

    S_H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    S_H_FILL  = PatternFill("solid", fgColor="2E75B6")
    S_L_FONT  = Font(name="Arial", bold=True, size=10)
    S_V_FONT  = Font(name="Arial", size=10)

    meta = [
        ["Field",          "Value"],
        ["Source URL",     BASE_URL],
        ["Pages Scraped",  f"{FIRST_PAGE} to {LAST_PAGE}  ({LAST_PAGE} pages)"],
        ["Total Agents",   len(df)],
        ["Columns",        ", ".join(FINAL_COLUMNS)],
        ["Scraped At",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Output File",    filepath],
    ]

    for r, row_vals in enumerate(meta, start=1):
        for c, val in enumerate(row_vals, start=1):
            cell            = summary.cell(row=r, column=c, value=val)
            cell.border     = BORDER
            cell.alignment  = Alignment(horizontal="left", vertical="center")
            if r == 1:
                cell.font = S_H_FONT
                cell.fill = S_H_FILL
            elif c == 1:
                cell.font = S_L_FONT
            else:
                cell.font = S_V_FONT

    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 50

    wb.save(filepath)
    log.info("Excel saved: %s  (%d agents, %d columns)", filepath, len(df), len(df.columns))


# ──────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    log.info("=" * 60)
    log.info("Kerala RERA Full Dataset Scraper")
    log.info("Pages  : %d  to  %d", FIRST_PAGE, LAST_PAGE)
    log.info("Output : %s", OUTPUT_FILE)
    log.info("=" * 60)

    # Phase 1 — Scrape all 45 pages
    raw_df = scrape_pages_1_to_45()

    if raw_df.empty:
        log.error(
            "No data collected. Possible causes:\n"
            "  1. %s is not reachable from your network.\n"
            "  2. The site has changed its URL or table structure.\n"
            "  3. Your IP may be rate-limited — wait a few minutes and retry.",
            BASE_URL,
        )
        sys.exit(1)

    # Phase 2 — Clean and normalize
    log.info("Cleaning data ...")
    clean_df = clean_dataframe(raw_df)

    if clean_df.empty:
        log.error("All rows removed during cleaning. Inspect the HTML table structure.")
        sys.exit(1)

    # Phase 3 — Export to Excel
    log.info("Writing Excel file ...")
    export_to_excel(clean_df, OUTPUT_FILE)

    log.info("All done!  Open '%s' to review your data.", OUTPUT_FILE)
    log.info("=" * 60)